#!/usr/bin/env python3
"""
Export FlexPrint order history and order-detail data to CSV (Excel-compatible).

Usage examples:
  FLEXPRINT_USER=Verbum FLEXPRINT_PASS=VerbuM23 python3 flexprint_order_export.py
  python3 flexprint_order_export.py --username Verbum --password 'VerbuM23' --view inprocess
  python3 flexprint_order_export.py --max-orders 20
"""

from __future__ import annotations

import argparse
import csv
import html
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import HTTPCookieProcessor, Request, build_opener as urllib_build_opener
import http.cookiejar


ORDER_COLUMNS = [
    "order_id",
    "created",
    "items",
    "total_price",
    "status",
    "group_id",
    "details_url",
]

DETAIL_COLUMNS = [
    "detail_created",
    "detail_subtotal",
    "detail_shipping",
    "detail_total_price",
    "detail_payment_received",
    "detail_balance_due",
    "detail_item_count",
    "detail_item_ids",
    "detail_item_descriptions",
    "detail_error",
]

ITEM_COLUMNS = [
    "order_id",
    "group_id",
    "details_url",
    "item_id",
    "description",
    "product",
    "quantity",
    "status",
    "comments",
    "destination",
    "price",
]


def normalize_base_url(url: str) -> str:
    url = url.strip()
    if not url.endswith("/"):
        url += "/"
    return url


def extract_hidden_inputs(page_html: str) -> dict[str, str]:
    hidden: dict[str, str] = {}
    for match in re.finditer(r'<input[^>]+type="hidden"[^>]*>', page_html, flags=re.I):
        tag = match.group(0)
        name_match = re.search(r'name="([^"]+)"', tag, flags=re.I)
        if not name_match:
            continue
        value_match = re.search(r'value="([^"]*)"', tag, flags=re.I)
        hidden[name_match.group(1)] = html.unescape(value_match.group(1) if value_match else "")
    return hidden


def strip_tags(raw_html: str) -> str:
    text = re.sub(r"<[^>]+>", " ", raw_html)
    text = html.unescape(text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_orders_table(page_html: str, base_url: str) -> list[dict[str, str]]:
    table_match = re.search(
        r'<table[^>]*id="OrdersDataGrid"[^>]*>.*?</table>',
        page_html,
        flags=re.I | re.S,
    )
    if not table_match:
        return []

    table_html = table_match.group(0)
    row_matches = list(re.finditer(r"<tr\b[^>]*>.*?</tr>", table_html, flags=re.I | re.S))
    if len(row_matches) <= 1:
        return []

    results: list[dict[str, str]] = []
    for row_match in row_matches[1:]:
        row_html = row_match.group(0)
        class_match = re.search(r'class="([^"]+)"', row_html, flags=re.I)
        class_tokens = class_match.group(1).split() if class_match else []
        group_id = next((token for token in class_tokens if token.isdigit()), "")

        cells = re.findall(r"<td\b[^>]*>(.*?)</td>", row_html, flags=re.I | re.S)
        if len(cells) < 6:
            continue

        results.append(
            {
                "order_id": strip_tags(cells[0]),
                "created": strip_tags(cells[1]),
                "items": strip_tags(cells[2]),
                "total_price": strip_tags(cells[3]),
                "status": strip_tags(cells[4]),
                "group_id": group_id,
                "details_url": f"{base_url}UserContentOrderSummary.aspx?group={group_id}&skipped=0" if group_id else "",
            }
        )
    return results


def parse_detail_header(page_html: str) -> dict[str, str]:
    header = {"detail_created": ""}
    match = re.search(
        r"Order\s+([^<\s]+)\s*<span[^>]*>.*?created\s+([^<]+)",
        page_html,
        flags=re.I | re.S,
    )
    if match:
        header["detail_created"] = strip_tags(match.group(2))
    return header


def parse_price_summary(page_html: str) -> dict[str, str]:
    result = {
        "detail_subtotal": "",
        "detail_shipping": "",
        "detail_total_price": "",
        "detail_payment_received": "",
        "detail_balance_due": "",
    }
    table_match = re.search(
        r"<table[^>]*id=['\"]priceSummary['\"][^>]*>.*?</table>",
        page_html,
        flags=re.I | re.S,
    )
    if not table_match:
        return result

    label_map = {
        "subtotal": "detail_subtotal",
        "shipping": "detail_shipping",
        "total price": "detail_total_price",
        "payment received": "detail_payment_received",
        "balance due": "detail_balance_due",
    }

    for row_match in re.finditer(r"<tr\b[^>]*>(.*?)</tr>", table_match.group(0), flags=re.I | re.S):
        cells = re.findall(r"<td\b[^>]*>(.*?)</td>", row_match.group(1), flags=re.I | re.S)
        if len(cells) < 2:
            continue
        label = strip_tags(cells[0]).lower()
        value = strip_tags(cells[-1])
        if label in label_map:
            result[label_map[label]] = value
    return result


def parse_detail_items(page_html: str) -> list[dict[str, str]]:
    table_match = re.search(
        r"<table[^>]*id=['\"]ShoppingCart\d+_ShoppingCart['\"][^>]*>.*?</table>",
        page_html,
        flags=re.I | re.S,
    )
    if not table_match:
        return []

    table_html = table_match.group(0)
    row_matches = list(re.finditer(r"<tr\b[^>]*>.*?</tr>", table_html, flags=re.I | re.S))
    if len(row_matches) <= 1:
        return []

    items: list[dict[str, str]] = []
    for row_match in row_matches[1:]:
        row_html = row_match.group(0)
        cells = re.findall(r"<td\b[^>]*>(.*?)</td>", row_html, flags=re.I | re.S)
        if len(cells) < 9:
            continue

        item_id_match = re.search(
            r"<p[^>]*class=['\"]itemTable-Id['\"][^>]*>.*?<span[^>]*>(.*?)</span>",
            cells[0],
            flags=re.I | re.S,
        )
        item_id = strip_tags(item_id_match.group(1)) if item_id_match else ""
        if not item_id:
            fallback = re.search(r"\b[A-Z]-[A-Z0-9]+\b", strip_tags(cells[0]))
            item_id = fallback.group(0) if fallback else ""

        items.append(
            {
                "item_id": item_id,
                "description": strip_tags(cells[1]),
                "product": strip_tags(cells[2]),
                "quantity": strip_tags(cells[3]),
                "status": strip_tags(cells[4]),
                "comments": strip_tags(cells[5]),
                "destination": strip_tags(cells[7]),
                "price": strip_tags(cells[8]),
            }
        )
    return items


def parse_order_detail(page_html: str) -> tuple[dict[str, str], list[dict[str, str]]]:
    summary: dict[str, str] = {}
    summary.update(parse_detail_header(page_html))
    summary.update(parse_price_summary(page_html))
    items = parse_detail_items(page_html)
    summary["detail_item_count"] = str(len(items))
    summary["detail_item_ids"] = " | ".join(item["item_id"] for item in items if item["item_id"])
    summary["detail_item_descriptions"] = " | ".join(item["description"] for item in items if item["description"])
    return summary, items


def create_http_opener():
    cookie_jar = http.cookiejar.CookieJar()
    opener = urllib_build_opener(HTTPCookieProcessor(cookie_jar))
    opener.addheaders = [("User-Agent", "Mozilla/5.0")]
    return opener


def get_page(opener, url: str, timeout: int = 60) -> tuple[str, str]:
    response = opener.open(url, timeout=timeout)
    return response.read().decode("utf-8", "ignore"), response.geturl()


def post_page(opener, url: str, payload: dict[str, str], timeout: int = 60) -> tuple[str, str]:
    request = Request(url, data=urlencode(payload).encode("utf-8"), method="POST")
    response = opener.open(request, timeout=timeout)
    return response.read().decode("utf-8", "ignore"), response.geturl()


def login(opener, base_url: str, username: str, password: str) -> None:
    login_url = f"{base_url}Login.aspx"
    login_page, _ = get_page(opener, login_url)
    hidden = extract_hidden_inputs(login_page)

    payload = {
        "__VIEWSTATE": hidden.get("__VIEWSTATE", ""),
        "__VIEWSTATEGENERATOR": hidden.get("__VIEWSTATEGENERATOR", ""),
        "__EVENTVALIDATION": hidden.get("__EVENTVALIDATION", ""),
        "Username": username,
        "Password": password,
        "SetAutoLogin": "on",
        "LoginToExistingAccount": "LoginToExistingAccount",
        "Cancel": "",
    }

    result_html, final_url = post_page(opener, login_url, payload)
    logged_in = ("Logout" in result_html) or final_url.lower().endswith("usercontentstart.aspx")
    if not logged_in:
        raise RuntimeError("Login failed. Check username/password.")


def set_orders_view(opener, base_url: str, page_html: str, view: str) -> str:
    if view == "all":
        return page_html

    orders_url = f"{base_url}UserContentOrders.aspx"
    hidden = extract_hidden_inputs(page_html)
    payload = {
        "__EVENTTARGET": "comboView",
        "__EVENTARGUMENT": "",
        "__VIEWSTATE": hidden.get("__VIEWSTATE", ""),
        "__VIEWSTATEGENERATOR": hidden.get("__VIEWSTATEGENERATOR", ""),
        "__EVENTVALIDATION": hidden.get("__EVENTVALIDATION", ""),
        "comboView": view,
    }
    page_html, _ = post_page(opener, orders_url, payload)
    return page_html


def write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(
    path: Path,
    orders_rows: list[dict[str, str]],
    items_rows: list[dict[str, str]],
    order_columns: list[str],
    item_columns: list[str],
) -> tuple[bool, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return (
            False,
            "openpyxl saknas i aktuell Python-runtime. Kör med bundled Python eller installera openpyxl.",
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Orders"
    items_sheet = workbook.create_sheet("Items")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def write_sheet(sheet, columns: list[str], rows: list[dict[str, str]]) -> None:
        sheet.append(columns)
        for row in rows:
            sheet.append([row.get(column, "") for column in columns])

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for col_idx, column_name in enumerate(columns, start=1):
            max_len = len(column_name)
            for row_idx in range(2, sheet.max_row + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                value_len = len(str(value)) if value is not None else 0
                if value_len > max_len:
                    max_len = value_len
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    write_sheet(orders_sheet, order_columns, orders_rows)
    write_sheet(items_sheet, item_columns, items_rows)
    workbook.save(path)
    return True, ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export FlexPrint order history to CSV.")
    parser.add_argument(
        "--base-url",
        default="https://live.flexprint.se/espressi/",
        help="Base URL to FlexPrint installation.",
    )
    parser.add_argument("--username", default=os.getenv("FLEXPRINT_USER"), help="FlexPrint username.")
    parser.add_argument("--password", default=os.getenv("FLEXPRINT_PASS"), help="FlexPrint password.")
    parser.add_argument(
        "--view",
        default="all",
        choices=["all", "pending", "inprocess", "completed", "unapproved"],
        help="Orders filter view.",
    )
    parser.add_argument(
        "--output",
        default=f"flexprint_orders_{os.getenv('USER', 'export')}.csv",
        help="Output CSV for order-level data.",
    )
    parser.add_argument(
        "--items-output",
        default=f"flexprint_order_items_{os.getenv('USER', 'export')}.csv",
        help="Output CSV for line-item data from each detail page.",
    )
    parser.add_argument(
        "--xlsx-output",
        default=f"flexprint_export_{os.getenv('USER', 'export')}.xlsx",
        help="Output XLSX workbook with two sheets: Orders and Items.",
    )
    parser.add_argument(
        "--max-orders",
        type=int,
        default=0,
        help="Optional limit. 0 means all orders.",
    )
    parser.add_argument(
        "--no-details",
        action="store_true",
        help="Skip loading each Details page.",
    )
    parser.add_argument(
        "--detail-timeout",
        type=int,
        default=90,
        help="Timeout in seconds per detail page request.",
    )
    parser.add_argument(
        "--detail-retries",
        type=int,
        default=3,
        help="Number of retries per detail page on timeout/network errors.",
    )
    parser.add_argument(
        "--detail-sleep",
        type=float,
        default=0.2,
        help="Sleep in seconds between detail page requests.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.username or not args.password:
        print("Missing credentials. Use --username/--password or FLEXPRINT_USER/FLEXPRINT_PASS.", file=sys.stderr)
        return 2

    base_url = normalize_base_url(args.base_url)
    opener = create_http_opener()

    try:
        login(opener, base_url, args.username, args.password)
        orders_url = f"{base_url}UserContentOrders.aspx"
        orders_html, _ = get_page(opener, orders_url)
        orders_html = set_orders_view(opener, base_url, orders_html, args.view)
        rows = parse_orders_table(orders_html, base_url)

        if args.max_orders > 0:
            rows = rows[: args.max_orders]

        items_export_rows: list[dict[str, str]] = []
        if not args.no_details:
            total = len(rows)
            for index, row in enumerate(rows, start=1):
                detail_url = row.get("details_url", "")
                if not detail_url:
                    row["detail_error"] = "No detail URL"
                    continue

                last_error = None
                detail_html = ""
                for attempt in range(1, max(1, args.detail_retries) + 1):
                    try:
                        detail_html, _ = get_page(opener, detail_url, timeout=max(5, args.detail_timeout))
                        last_error = None
                        break
                    except Exception as exc:  # noqa: BLE001
                        last_error = exc
                        if attempt < max(1, args.detail_retries):
                            time.sleep(min(2.0 * attempt, 5.0))

                if last_error is not None:
                    row["detail_error"] = str(last_error)
                else:
                    detail_summary, detail_items = parse_order_detail(detail_html)
                    row.update(detail_summary)
                    row["detail_error"] = ""

                    for item in detail_items:
                        item_row = {
                            "order_id": row.get("order_id", ""),
                            "group_id": row.get("group_id", ""),
                            "details_url": detail_url,
                        }
                        item_row.update(item)
                        items_export_rows.append(item_row)

                if args.detail_sleep > 0:
                    time.sleep(args.detail_sleep)

                if index % 10 == 0 or index == total:
                    print(f"Fetched details {index}/{total}")

        write_csv(Path(args.output), rows, ORDER_COLUMNS + DETAIL_COLUMNS)
        if not args.no_details:
            write_csv(Path(args.items_output), items_export_rows, ITEM_COLUMNS)
        workbook_written = False
        workbook_message = ""
        if args.xlsx_output:
            workbook_written, workbook_message = write_xlsx(
                path=Path(args.xlsx_output),
                orders_rows=rows,
                items_rows=items_export_rows,
                order_columns=ORDER_COLUMNS + DETAIL_COLUMNS,
                item_columns=ITEM_COLUMNS,
            )

    except Exception as exc:  # noqa: BLE001
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    if args.no_details:
        print(f"Done: {len(rows)} orders exported to {Path(args.output).resolve()}")
    else:
        print(
            "Done: "
            f"{len(rows)} orders -> {Path(args.output).resolve()} and "
            f"{Path(args.items_output).resolve()}"
        )
    if args.xlsx_output and workbook_written:
        print(f"Workbook: {Path(args.xlsx_output).resolve()}")
    elif args.xlsx_output and not workbook_written:
        print(f"Workbook skipped: {workbook_message}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
